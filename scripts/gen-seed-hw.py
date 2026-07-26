# -*- coding: utf-8 -*-
"""
Genera worker/seed-hw.sql a partir de INVENTARIO_HW.xlsx (hoja 'Inventario').
Cada fila -> una fila en ig_queue con fuente='drive':
  - ml_item_id = 'hw-<N°>'  (sintético, único; no choca con ids de ML)
  - titulo     = 'Marca Modelo'  (para los avisos de Telegram)
  - image_url  = <PUBLIC_URL>/img/drive/<fileId de Drive>
  - caption    = ficha del auto + disponibilidad + "precio por DM"
  - precio=0, estado='pendiente'
Uso:  python scripts/gen-seed-hw.py [ruta_xlsx] [salida.sql]
"""
import openpyxl, re, sys, os

XLSX = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\aronr\OneDrive\Escritorio\INVENTARIO_HW.xlsx"
OUT  = sys.argv[2] if len(sys.argv) > 2 else os.path.join(os.path.dirname(__file__), "..", "worker", "seed-hw.sql")
PUBLIC_URL = "https://mlpu-proxy.aronricocl.workers.dev"
HASHTAGS = "#hotwheels #matchbox #diecast #hotwheelschile #coleccionables #autosaescala #chile"

def sq(s):  # escapa para string SQL de SQLite
    return "'" + str(s).replace("'", "''") + "'"

def drive_id(url):
    if not url:
        return None
    m = re.search(r"/file/d/([A-Za-z0-9_-]{10,})", url) or re.search(r"[?&]id=([A-Za-z0-9_-]{10,})", url)
    return m.group(1) if m else None

def clean(v):
    return str(v).strip() if v not in (None, "") else ""

def build_caption(marca, modelo, serie, color, ncol, disp):
    lineas = [f"🏎️ {marca} {modelo}".rstrip(), ""]
    ficha = []
    if serie: ficha.append(f"🏁 Serie: {serie}")
    if color: ficha.append(f"🎨 Color: {color}")
    if ncol:  ficha.append(f"🔢 Colección: {ncol}")
    if ficha:
        lineas += ficha + [""]
    disponible = "disponible" in disp.lower() or "✅" in disp
    lineas.append("🟢 DISPONIBLE" if disponible else "🔴 AGOTADO")
    lineas.append("💬 Precio: consúltalo por DM 📩")
    lineas += ["", "📦 Envíos a todo Chile", "👉 Escríbenos por DM para comprar", "", HASHTAGS]
    return "\n".join(lineas)

wb = openpyxl.load_workbook(XLSX)
ws = wb["Inventario"]

filas, sin_foto = [], []
for r in range(2, ws.max_row + 1):
    n     = ws.cell(r, 1).value
    marca = clean(ws.cell(r, 2).value)
    modelo= clean(ws.cell(r, 3).value)
    serie = clean(ws.cell(r, 4).value)
    color = clean(ws.cell(r, 5).value)
    ncol  = clean(ws.cell(r, 6).value)
    disp  = clean(ws.cell(r, 8).value)
    link_cell = ws.cell(r, 11)
    target = link_cell.hyperlink.target if link_cell.hyperlink else None
    fid = drive_id(target)
    if n is None or not modelo:
        continue
    if not fid:
        sin_foto.append((n, marca, modelo))
        continue
    titulo  = f"{marca} {modelo}".strip()
    caption = build_caption(marca, modelo, serie, color, ncol, disp)
    image   = f"{PUBLIC_URL}/img/drive/{fid}"
    filas.append((f"hw-{n}", titulo, image, caption))

lines = [
    "-- seed-hw.sql — inventario propio (Google Drive) para publicar en Instagram.",
    "-- Generado por scripts/gen-seed-hw.py. Requiere antes: migrate-drive.sql aplicado.",
    "-- Aplicar (esto dispara la publicación en el próximo tick del cron / con /ig rush):",
    "--   npx wrangler d1 execute mlpu-db --remote --file=worker/seed-hw.sql",
    "",
]
for mlid, titulo, image, caption in filas:
    lines.append(
        "INSERT OR IGNORE INTO ig_queue (ml_item_id, titulo, precio, estado, fuente, image_url, caption) VALUES ("
        f"{sq(mlid)}, {sq(titulo)}, 0, 'pendiente', 'drive', {sq(image)}, {sq(caption)});"
    )

with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(lines) + "\n")

print(f"OK -> {os.path.normpath(OUT)}")
print(f"Filas generadas: {len(filas)}")
if sin_foto:
    print(f"Sin foto de Drive (omitidas): {len(sin_foto)} -> {sin_foto[:5]}")
print("\n--- CAPTION DE EJEMPLO (fila 1) ---")
if filas:
    print(filas[0][3])
